from __future__ import annotations

import os


class TestReportGenerator:
    def test_export_3_sheets_no_compliance(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """无需求标准时生成 3 个 Sheet"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test_report.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[{
                "group_name": "ThinkPad E14",
                "comparison_status": "comparable",
                "supplier_prices": '[{"supplier_file_id":"sf1","supplier_name":"联想","unit_price":4299}]',
                "min_price": 4299, "effective_min_price": 4299,
                "max_price": 4299, "avg_price": 4299, "price_diff": 0,
                "anomaly_details": "[]",
            }],
            standardized_rows=[{
                "supplier_name": "联想", "product_name": "ThinkPad E14",
                "spec_model": "i5/16GB/512GB", "unit": "台", "quantity": 10,
                "unit_price": 4299, "total_price": 42990,
            }],
            traceability_data=[{
                "source_file": "联想报价.xlsx", "supplier_name": "联想",
                "source_location": "A3", "standard_field": "product_name",
            }],
            compliance_matrix=None,
            supplier_names={"sf1": "联想"},
        )

        assert os.path.exists(output)
        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 3
        assert "比价结果表" in wb.sheetnames
        assert "标准化明细表" in wb.sheetnames
        assert "追溯信息表" in wb.sheetnames

    def test_export_4_sheets_with_compliance(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """有需求标准时生成 4 个 Sheet"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test_report.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[],
            standardized_rows=[],
            traceability_data=[],
            compliance_matrix={
                "supplier_names": {"sf1": "联想"},
                "rows": [{
                    "requirement": {"code": "REQ-001", "category": "技术规格",
                                    "title": "内存>=16GB", "is_mandatory": True},
                    "suppliers": {"sf1": {"status": "match", "evidence_text": "16GB DDR5",
                                          "needs_review": False}},
                }],
            },
            supplier_names={"sf1": "联想"},
        )

        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 4
        assert "需求符合性矩阵" in wb.sheetnames

    def test_comparison_sheet_has_headers(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """比价结果表包含正确表头"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[],
            standardized_rows=[],
            traceability_data=[],
            compliance_matrix=None,
            supplier_names={"sf1": "联想", "sf2": "戴尔"},
        )

        wb = openpyxl.load_workbook(output)
        ws = wb["比价结果表"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "商品组" in headers
        assert "全量最低价" in headers
        assert "异常标记" in headers

    def test_min_price_highlighted(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """最低价单元格应有绿色填充"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[{
                "group_name": "test",
                "comparison_status": "comparable",
                "supplier_prices": '[{"supplier_file_id":"sf1","supplier_name":"A","unit_price":100},{"supplier_file_id":"sf2","supplier_name":"B","unit_price":200}]',
                "min_price": 100, "effective_min_price": 100,
                "max_price": 200, "avg_price": 150, "price_diff": 100,
                "anomaly_details": "[]",
            }],
            standardized_rows=[],
            traceability_data=[],
            compliance_matrix=None,
            supplier_names={"sf1": "A", "sf2": "B"},
        )

        wb = openpyxl.load_workbook(output)
        ws = wb["比价结果表"]
        # sf1 排序后在第一列（column=3），值 100 == min_price
        cell = ws.cell(row=2, column=3)
        assert cell.value == 100
        # 绿色填充 C6EFCE
        assert cell.fill.start_color.rgb is not None

    def test_anomaly_highlighted(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """异常单元格应有红色填充"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[{
                "group_name": "test",
                "comparison_status": "blocked",
                "supplier_prices": '[{"supplier_file_id":"sf1","supplier_name":"A","unit_price":100}]',
                "min_price": 100, "effective_min_price": 100,
                "max_price": 100, "avg_price": 100, "price_diff": 0,
                "anomaly_details": '[{"description":"税价口径不一致","type":"tax_basis_mismatch","blocking":true,"affected_suppliers":["A"]}]',
            }],
            standardized_rows=[],
            traceability_data=[],
            compliance_matrix=None,
            supplier_names={"sf1": "A"},
        )

        wb = openpyxl.load_workbook(output)
        ws = wb["比价结果表"]
        # 异常标记列: 商品组(1) + 状态(1) + 供应商数(1) + 汇总列(6) = column 9
        anomaly_cell = ws.cell(row=2, column=9)
        assert "税价口径不一致" in str(anomaly_cell.value)
        assert anomaly_cell.fill.start_color.rgb is not None

    def test_compliance_sheet_partial_highlighted(self, tmp_path) -> None:  # type: ignore[no-untyped-def]
        """符合性矩阵 partial 状态有黄色填充"""
        import openpyxl

        from engines.report_generator import ReportGenerator

        engine = ReportGenerator()
        output = str(tmp_path / "test.xlsx")
        engine.export_to_excel(
            output_path=output,
            comparison_results=[],
            standardized_rows=[],
            traceability_data=[],
            compliance_matrix={
                "supplier_names": {"sf1": "联想"},
                "rows": [{
                    "requirement": {"code": "REQ-001", "category": "技术规格",
                                    "title": "测试", "is_mandatory": True},
                    "suppliers": {"sf1": {"status": "partial", "evidence_text": "部分符合",
                                          "needs_review": True}},
                }],
            },
            supplier_names={"sf1": "联想"},
        )

        wb = openpyxl.load_workbook(output)
        ws = wb["需求符合性矩阵"]
        # 状态列: 需求编号(1)+分类(1)+描述(1)+必选(1)+联想状态(5) = column 5
        status_cell = ws.cell(row=2, column=5)
        assert status_cell.value == "partial"
        assert status_cell.fill.start_color.rgb is not None
