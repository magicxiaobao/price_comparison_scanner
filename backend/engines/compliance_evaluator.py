from __future__ import annotations

import re
from dataclasses import dataclass


@dataclass
class ParsedRequirement:
    """从 Excel 模板解析出的需求项"""

    code: str | None
    category: str
    title: str
    description: str | None
    is_mandatory: bool
    match_type: str
    expected_value: str | None
    operator: str | None


@dataclass
class MatchResult:
    """单条匹配结果"""

    status: str  # match / partial / no_match / unclear
    match_score: float  # 0.0 - 1.0
    evidence_text: str
    evidence_location: str  # JSON 字符串
    match_method: str  # keyword / numeric / manual
    needs_review: bool


class ComplianceEvaluator:
    """需求标准管理 + 供应商符合性匹配引擎（可选模块）"""

    ENGINE_VERSION = "compliance_evaluator:1.0"

    # ---- 符合性匹配 ----

    def evaluate_single(
        self,
        requirement: dict,
        supplier_rows: list[dict],
        supplier_file_id: str,
    ) -> MatchResult:
        """对单个「需求项 x 供应商（该商品组下的行）」执行匹配。"""
        match_type = requirement["match_type"]

        if match_type == "keyword":
            return self._match_keyword(requirement, supplier_rows)
        elif match_type == "numeric":
            return self._match_numeric(requirement, supplier_rows)
        else:  # manual
            return MatchResult(
                status="unclear",
                match_score=0.0,
                evidence_text="",
                evidence_location="{}",
                match_method="manual",
                needs_review=True,
            )

    def _match_keyword(
        self, requirement: dict, rows: list[dict]
    ) -> MatchResult:
        """keyword 匹配：在 spec_model / remark / product_name 中搜索关键词。"""
        keyword = (requirement.get("expected_value") or "").strip()
        if not keyword:
            return MatchResult(
                status="unclear",
                match_score=0.0,
                evidence_text="未设置关键词",
                evidence_location="{}",
                match_method="keyword",
                needs_review=True,
            )

        search_fields = ["spec_model", "remark", "product_name"]
        for row in rows:
            for field_name in search_fields:
                field_value = str(row.get(field_name, "") or "")
                if keyword.lower() in field_value.lower():
                    return MatchResult(
                        status="match",
                        match_score=1.0,
                        evidence_text=f"在 {field_name} 中找到关键词「{keyword}」: {field_value}",
                        evidence_location="{}",
                        match_method="keyword",
                        needs_review=False,
                    )

        return MatchResult(
            status="unclear",
            match_score=0.0,
            evidence_text=f"未在供应商数据中找到关键词「{keyword}」",
            evidence_location="{}",
            match_method="keyword",
            needs_review=True,
        )

    def _match_numeric(
        self, requirement: dict, rows: list[dict]
    ) -> MatchResult:
        """numeric 匹配：提取数值与目标值比较。"""
        expected_str = (requirement.get("expected_value") or "").strip()
        operator = requirement.get("operator", "gte")

        try:
            expected = float(expected_str)
        except (ValueError, TypeError):
            return MatchResult(
                status="unclear",
                match_score=0.0,
                evidence_text=f"无法解析目标值: {expected_str}",
                evidence_location="{}",
                match_method="numeric",
                needs_review=True,
            )

        search_fields = ["spec_model", "remark", "product_name"]
        for row in rows:
            for field_name in search_fields:
                field_value = str(row.get(field_name, "") or "")
                numbers = re.findall(r"[\d]+\.?\d*", field_value)
                for num_str in numbers:
                    try:
                        actual = float(num_str)
                    except ValueError:
                        continue

                    satisfied = self._compare_numeric(actual, expected, operator)
                    if satisfied is not None:
                        status = "match" if satisfied else "no_match"
                        return MatchResult(
                            status=status,
                            match_score=1.0 if satisfied else 0.0,
                            evidence_text=(
                                f"在 {field_name} 中提取数值 {actual}，"
                                f"目标 {operator} {expected}"
                            ),
                            evidence_location="{}",
                            match_method="numeric",
                            needs_review=False,
                        )

        return MatchResult(
            status="unclear",
            match_score=0.0,
            evidence_text="未能从供应商数据中提取可比较的数值",
            evidence_location="{}",
            match_method="numeric",
            needs_review=True,
        )

    def _compare_numeric(
        self, actual: float, expected: float, operator: str
    ) -> bool | None:
        if operator == "gte":
            return actual >= expected
        elif operator == "lte":
            return actual <= expected
        elif operator == "eq":
            return abs(actual - expected) < 0.001
        elif operator == "range":
            return None  # range 需要两个值，MVP 简化处理
        return None

    # ---- 导入解析 ----

    def parse_requirements_excel(self, file_path: str) -> list[ParsedRequirement]:
        """从模板 Excel 导入需求标准。

        模板格式：
        | 需求编号 | 需求分类 | 需求标题 | 需求描述 | 是否必选 | 判断类型 | 目标值 | 比较操作符 |
        """
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        results: list[ParsedRequirement] = []

        header_row = 1
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[2]:  # title 列为空则跳过
                continue
            code = str(row[0]).strip() if row[0] else None
            category = str(row[1]).strip() if row[1] else "功能要求"
            title = str(row[2]).strip()
            description = str(row[3]).strip() if row[3] else None
            is_mandatory = (
                str(row[4]).strip() in ("是", "必选", "1", "true", "True")
                if row[4]
                else True
            )
            match_type = str(row[5]).strip().lower() if row[5] else "manual"
            if match_type not in ("keyword", "numeric", "manual"):
                match_type = "manual"
            expected_value = str(row[6]).strip() if row[6] else None
            operator = (
                str(row[7]).strip().lower() if len(row) > 7 and row[7] else None
            )
            if operator and operator not in ("gte", "lte", "eq", "range"):
                operator = None

            results.append(
                ParsedRequirement(
                    code=code,
                    category=(
                        category
                        if category
                        in ("功能要求", "技术规格", "商务条款", "服务要求", "交付要求")
                        else "功能要求"
                    ),
                    title=title,
                    description=description,
                    is_mandatory=is_mandatory,
                    match_type=match_type,
                    expected_value=expected_value,
                    operator=operator,
                )
            )

        wb.close()
        return results

    def export_requirements_template(
        self, requirements: list[dict], output_path: str
    ) -> str:
        """导出需求标准为 Excel 模板。"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "需求标准"

        headers = [
            "需求编号", "需求分类", "需求标题", "需求描述",
            "是否必选", "判断类型", "目标值", "比较操作符",
        ]
        ws.append(headers)

        for req in requirements:
            ws.append([
                req.get("code", ""),
                req.get("category", ""),
                req.get("title", ""),
                req.get("description", ""),
                "是" if req.get("is_mandatory", True) else "否",
                req.get("match_type", ""),
                req.get("expected_value", ""),
                req.get("operator", ""),
            ])

        wb.save(output_path)
        return output_path

    # ---- 匹配逻辑在 Task 4.2 中实现 ----
