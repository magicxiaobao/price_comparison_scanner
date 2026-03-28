from __future__ import annotations

import contextlib
import json
from pathlib import Path


class ReportGenerator:
    """4 Sheet Excel 导出引擎"""

    ENGINE_VERSION = "report_generator:1.0"

    def export_to_excel(
        self,
        output_path: str,
        comparison_results: list[dict],
        standardized_rows: list[dict],
        traceability_data: list[dict],
        compliance_matrix: dict | None,
        supplier_names: dict[str, str],
    ) -> str:
        """生成 Excel 审计底稿（3 或 4 Sheet）。"""
        import openpyxl
        from openpyxl.styles import Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        min_price_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        effective_min_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        anomaly_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "比价结果表"
        self._write_comparison_sheet(
            ws1, comparison_results, supplier_names,
            compliance_matrix is not None,
            header_font, header_fill, min_price_fill,
            effective_min_fill, anomaly_fill, thin_border,
        )

        ws2 = wb.create_sheet("标准化明细表")
        self._write_standardized_sheet(
            ws2, standardized_rows, header_font, header_fill, thin_border,
        )

        ws3 = wb.create_sheet("追溯信息表")
        self._write_traceability_sheet(
            ws3, traceability_data, header_font, header_fill, thin_border,
        )

        if compliance_matrix is not None:
            ws4 = wb.create_sheet("需求符合性矩阵")
            self._write_compliance_sheet(
                ws4, compliance_matrix, supplier_names,
                header_font, header_fill, partial_fill, thin_border,
            )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    # ---- Sheet writers ----

    def _write_comparison_sheet(  # noqa: PLR0913
        self, ws, results, supplier_names, has_compliance,  # type: ignore[no-untyped-def]
        header_font, header_fill, min_fill, eff_fill, anomaly_fill, border,
    ) -> None:
        sorted_sids = sorted(supplier_names.keys())
        headers = ["商品组", "比较状态"]
        for sid in sorted_sids:
            headers.append(f"{supplier_names[sid]} 单价")
        headers.extend(["全量最低价", "有效最低价", "最高价", "平均价", "差额", "异常标记"])
        if has_compliance:
            headers.append("符合性摘要")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r.get("group_name", "")).border = border
            ws.cell(row=row_idx, column=2, value=r.get("comparison_status", "")).border = border

            raw_sp = r.get("supplier_prices", "[]")
            supplier_prices = json.loads(raw_sp) if isinstance(raw_sp, str) else raw_sp
            sp_map = {sp["supplier_file_id"]: sp for sp in supplier_prices}

            for col_offset, sid in enumerate(sorted_sids):
                sp = sp_map.get(sid, {})
                up = sp.get("unit_price")
                cell = ws.cell(row=row_idx, column=3 + col_offset, value=up)
                cell.border = border
                if up is not None and up == r.get("min_price"):
                    cell.fill = min_fill
                if (has_compliance and up is not None
                        and up == r.get("effective_min_price")
                        and r.get("effective_min_price") != r.get("min_price")):
                    cell.fill = eff_fill

            base_col = 3 + len(sorted_sids)
            ws.cell(row=row_idx, column=base_col, value=r.get("min_price")).border = border
            ws.cell(row=row_idx, column=base_col + 1, value=r.get("effective_min_price")).border = border
            ws.cell(row=row_idx, column=base_col + 2, value=r.get("max_price")).border = border
            ws.cell(row=row_idx, column=base_col + 3, value=r.get("avg_price")).border = border
            ws.cell(row=row_idx, column=base_col + 4, value=r.get("price_diff")).border = border

            raw_ad = r.get("anomaly_details", "[]")
            anomalies = json.loads(raw_ad) if isinstance(raw_ad, str) else raw_ad
            anomaly_text = "; ".join(a["description"] for a in anomalies) if anomalies else ""
            anomaly_cell = ws.cell(row=row_idx, column=base_col + 5, value=anomaly_text)
            anomaly_cell.border = border
            if anomalies:
                anomaly_cell.fill = anomaly_fill

    def _write_standardized_sheet(  # type: ignore[no-untyped-def]
        self, ws, rows, header_font, header_fill, border,
    ) -> None:
        headers = [
            "供应商", "商品名称", "规格型号", "单位", "数量", "单价", "总价",
            "税率", "税价口径", "备注", "是否人工修改",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=r.get("supplier_name", "")).border = border
            ws.cell(row=row_idx, column=2, value=r.get("product_name", "")).border = border
            ws.cell(row=row_idx, column=3, value=r.get("spec_model", "")).border = border
            ws.cell(row=row_idx, column=4, value=r.get("unit", "")).border = border
            ws.cell(row=row_idx, column=5, value=r.get("quantity")).border = border
            ws.cell(row=row_idx, column=6, value=r.get("unit_price")).border = border
            ws.cell(row=row_idx, column=7, value=r.get("total_price")).border = border
            ws.cell(row=row_idx, column=8, value=r.get("tax_rate")).border = border
            ws.cell(row=row_idx, column=9, value=r.get("tax_basis", "")).border = border
            ws.cell(row=row_idx, column=10, value=r.get("remark", "")).border = border
            ws.cell(
                row=row_idx, column=11,
                value="是" if r.get("is_manually_modified") else "否",
            ).border = border

    def _write_traceability_sheet(  # type: ignore[no-untyped-def]
        self, ws, data, header_font, header_fill, border,
    ) -> None:
        headers = [
            "来源文件名", "供应商", "来源定位", "命中规则",
            "置信度", "是否需复核", "是否人工修改",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, r in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=r.get("source_file", "")).border = border
            ws.cell(row=row_idx, column=2, value=r.get("supplier_name", "")).border = border

            source_loc = r.get("source_location", "")
            if isinstance(source_loc, str) and source_loc.startswith("{"):
                with contextlib.suppress(json.JSONDecodeError, TypeError):
                    source_loc = json.dumps(json.loads(source_loc), ensure_ascii=False)
            ws.cell(row=row_idx, column=3, value=str(source_loc)).border = border

            hit_rules = r.get("hit_rule_snapshots", "")
            if isinstance(hit_rules, str) and hit_rules.startswith("["):
                try:
                    rules = json.loads(hit_rules)
                    hit_rules = "; ".join(
                        rule.get("rule_name", "") for rule in rules if rule.get("rule_name")
                    )
                except (json.JSONDecodeError, TypeError):
                    pass
            ws.cell(row=row_idx, column=4, value=str(hit_rules)).border = border

            ws.cell(row=row_idx, column=5, value=r.get("confidence")).border = border
            ws.cell(
                row=row_idx, column=6,
                value="是" if r.get("needs_review") else "否",
            ).border = border
            ws.cell(
                row=row_idx, column=7,
                value="是" if r.get("is_manually_modified") else "否",
            ).border = border

    def _write_compliance_sheet(  # type: ignore[no-untyped-def]
        self, ws, matrix, supplier_names, header_font, header_fill, partial_fill, border,
    ) -> None:
        sorted_sids = sorted(supplier_names.keys())

        base_headers = ["需求编号", "需求分类", "需求描述", "是否必选"]
        for sid in sorted_sids:
            name = supplier_names[sid]
            base_headers.append(f"{name} 状态")
            base_headers.append(f"{name} 证据摘要")
        base_headers.append("是否人工确认")

        for col_idx, h in enumerate(base_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        rows = matrix.get("rows", [])
        for row_idx, mr in enumerate(rows, 2):
            req = mr.get("requirement", {})
            ws.cell(row=row_idx, column=1, value=req.get("code", "")).border = border
            ws.cell(row=row_idx, column=2, value=req.get("category", "")).border = border
            ws.cell(row=row_idx, column=3, value=req.get("title", "")).border = border
            ws.cell(
                row=row_idx, column=4,
                value="是" if req.get("is_mandatory") else "否",
            ).border = border

            suppliers = mr.get("suppliers", {})
            col_offset = 5
            any_confirmed = False
            for sid in sorted_sids:
                cell_data = suppliers.get(sid, {})
                status = cell_data.get("status", "")
                evidence = cell_data.get("evidence_text", "")
                status_cell = ws.cell(row=row_idx, column=col_offset, value=status)
                status_cell.border = border
                if status == "partial":
                    status_cell.fill = partial_fill
                ws.cell(row=row_idx, column=col_offset + 1, value=evidence or "").border = border
                col_offset += 2
                if not cell_data.get("needs_review", True):
                    any_confirmed = True

            ws.cell(
                row=row_idx, column=col_offset,
                value="是" if any_confirmed else "否",
            ).border = border
