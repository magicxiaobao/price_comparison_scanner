from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path


@dataclass
class RawTableData:
    """解析器内部使用的原始表格数据结构"""

    table_index: int
    headers: list[str]  # 表头行
    rows: list[list[str | None]]  # 数据行（每行为单元格值列表）
    sheet_name: str | None = None  # Excel sheet 名
    page_number: int | None = None  # PDF 页码
    row_count: int = 0
    column_count: int = 0


class DocumentParser:
    """
    根据文件类型自动分发到对应解析器。
    返回 RawTableData 列表。
    """

    SUPPORTED_TYPES = {"xlsx", "docx", "pdf"}

    def parse(
        self, file_path: str, progress_callback: Callable[[float], None] | None = None
    ) -> list[RawTableData]:
        """
        解析文件，返回 RawTableData 列表。
        progress_callback: 可选，用于报告进度（0.0-1.0）。
        """
        path = Path(file_path)
        suffix = path.suffix.lower().lstrip(".")

        if suffix == "xlsx":
            return self._parse_xlsx(file_path, progress_callback)
        elif suffix == "docx":
            return self._parse_docx(file_path, progress_callback)
        elif suffix == "pdf":
            return self._parse_pdf(file_path, progress_callback)
        else:
            raise ValueError(f"不支持的文件类型: {suffix}")

    def _parse_xlsx(
        self, file_path: str, progress_callback: Callable[[float], None] | None = None
    ) -> list[RawTableData]:
        """
        使用 openpyxl 读取 Excel 文件。
        每个 sheet 生成一个 RawTableData。
        跳过完全空白的 sheet。
        """
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        results: list[RawTableData] = []
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)

        for idx, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name]

            all_rows: list[list[str | None]] = []
            for row in ws.iter_rows():
                cell_values = [
                    str(cell.value).strip() if cell.value is not None else None
                    for cell in row
                ]
                all_rows.append(cell_values)

            # 跳过完全空白的 sheet
            has_data = any(any(v is not None for v in row) for row in all_rows)
            if not has_data:
                if progress_callback:
                    progress_callback((idx + 1) / total_sheets)
                continue

            # 第一行作为表头
            headers = [v or "" for v in all_rows[0]] if all_rows else []
            data_rows = all_rows[1:] if len(all_rows) > 1 else []

            table = RawTableData(
                table_index=len(results),
                headers=headers,
                rows=data_rows,
                sheet_name=sheet_name,
                row_count=len(data_rows),
                column_count=len(headers),
            )
            results.append(table)

            if progress_callback:
                progress_callback((idx + 1) / total_sheets)

        wb.close()
        return results

    def _parse_docx(
        self, file_path: str, progress_callback: Callable[[float], None] | None = None
    ) -> list[RawTableData]:
        """Word 解析器占位 — Task 1.3 实现"""
        raise NotImplementedError("Word 解析器尚未实现")

    def _parse_pdf(
        self, file_path: str, progress_callback: Callable[[float], None] | None = None
    ) -> list[RawTableData]:
        """PDF 解析器占位 — Task 1.4 实现"""
        raise NotImplementedError("PDF 解析器尚未实现")

    @staticmethod
    def _is_ocr_available() -> bool:
        """检测 OCR 模块（PaddleOCR）是否已安装"""
        try:
            import paddleocr  # noqa: F401

            return True
        except ImportError:
            return False

    def _fallback_ocr(
        self, file_path: str, progress_callback: Callable[[float], None] | None = None
    ) -> list[RawTableData]:
        """OCR 降级解析 — Phase 5 实现，当前返回空列表 + 提示"""
        if not self._is_ocr_available():
            return []
        raise NotImplementedError("OCR 解析尚未实现")
